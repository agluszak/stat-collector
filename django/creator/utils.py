import io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def create_xls(data: dict):
    wb = Workbook()
    ws = wb.active
    ws.title = "stats"

    headers = (
        "client",
        "supplier",
        "placementType",
        "copy",
        "statName",
        "startDate",
        "endDate",
        "statVal",
    )
    ws.append(headers)
    name = data.get("name", "Statistics")
    client = data.get("client")
    periods = data.get("periods")
    placements = data.get("placementTypes")

    for placement in placements:
        placement_name = placement.get("name")
        suppliers = placement.get("suppliers")
        statistics = placement.get("statistics")
        copies = placement.get("copies")
        for supplier in suppliers:
            supplier_name = supplier.get("name")
            stats = supplier.get("stats")
            for period_idx, period in enumerate(periods):
                for statistic_idx, statistic in enumerate(statistics):
                    for copy_idx, copy in enumerate(copies):
                        start_date = period.get("startDate").replace(".", "-")
                        end_date = period.get("endDate").replace(".", "-")
                        row = (
                            client,
                            supplier_name,
                            placement_name,
                            copy,
                            statistic,
                            start_date,
                            end_date,
                            stats[statistic_idx][copy_idx][period_idx],
                        )
                        ws.append(row)

    # XLS beautification

    # Freeze panes
    ws.freeze_panes = ws["A2"]

    # Adjust the width of the columns
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return name, output.getvalue()
